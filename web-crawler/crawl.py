import re
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook, cell


def get_start_url(listing):
    if listing.endswith('/'):
        if listing.startswith('w'):
            listings = "http://" + listing + where.capitalize() + '/' + what.capitalize()
        else:
            listings = listing + where.capitalize() + '/' + what.capitalize()
    else:
        listings = "http://" + listing + '/' + where.capitalize() + '/' + what.capitalize()
    return listings


def get_html(url):
    """

    :param url:
    :return:
    """
    message = requests.get(url)
    return message


def get_next_link(msg):
    class_regex = re.compile('<div\sclass=[\'"]pagination[\'"]>(.*?)</div>')
    a_list = class_regex.findall(msg)
    if len(a_list) > 0:
        a = a_list[0].split('</a>')
        a = [i for i in a if 'next' in i]
        link_regex = re.compile('<a\shref=[\'"](.*?)[\'"]>next\s')
        next_link = link_regex.findall(a[0])
        if len(next_link) > 0:
            return next_link[0]
        else:
            return ''
    else:
        return ''


def get_company_name(soup):
    """

    :param soup:
    :return:
    """

    c_name = soup.find_all('a', "jrat")
    cname = []
    for i in c_name:
        cname.append(i.text)
    return cname


def get_phone(soup):
    divtag = soup.findAll("ctc f13")
    phone = [i.p.text.replace('Call:', '') for i in divtag]
    return phone


def get_add(soup):
    add_list = soup.findAll("jrc13")
    add = [i.text.split('|')[0].replace('\t', '') for i in add_list]
    return add


if __name__ == '__main__':
    listing = "www.justdial.com"
    where = input("City: ")
    what = input("Category: ")
    url = get_start_url(listing)

    name = []
    person = []
    phone = []
    add = []

    while url:
        if ' ' in url:
            url = url.replace(' ', '%20')
            #print(url)
        msg = (requests.get(url)).text
        print(url)
        soup = BeautifulSoup(msg)
        print(soup)
        for i in get_company_name(soup):
        #for tag in soup.find_all('a', 'jrat'):
            name.append(i)
        for i in get_phone(soup):
            phone.append(i)
        for i in get_add(soup):
            add.append(i)
        url = get_next_link(msg)
    book = Workbook()
    sheet1 = book.active
    index = ['NAME', 'PHONE', 'ADDRESS']

    # style = xlwt.XFStyle()
    #font = xlwt.Font()
    #font.name = 'Times New Roman'
    #font.bold = True
    #style.font = font

    # for n in range(0,3):
    # sheet1.write(0,n,index[n].upper(),style)

    for i in range(0, len(name)):
        sheet1.cell('%s%s' % (i + 1, 0)).value = name[i]
        sheet1.cell('%s%s' % (i + 1, 1)).value = phone[i]
        sheet1.cell('%s%s' % (i + 1, 2)).value = add[i]

    book.save(what + '.xlsx')
